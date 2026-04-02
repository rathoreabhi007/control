"""
Config Validator Script
Runs as a subprocess to validate Excel config files.
Reads the Excel, validates each sheet, and writes results JSON + validated Excel.

Environment variables:
    TASK_ID: Unique task identifier
    FILE_PATH: Path to the Excel config file
    RESULTS_DIR: Directory to write results JSON
    TASK_FILE_PATH: Path to task state file (set by control_runner)
"""
import os
import sys
import json
import traceback
from pathlib import Path
from datetime import datetime
from copy import copy

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def get_env(name, default=None):
    value = os.environ.get(name, default)
    if value is None:
        print(f"WARNING: Environment variable {name} not set")
    return value


def read_excel_sheets(file_path):
    """Read all sheets from an Excel file using openpyxl. Returns dict of sheet_name -> list of row dicts."""
    print(f"Reading Excel file: {file_path}")
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    print(f"Found {len(sheet_names)} sheets: {sheet_names}")

    sheets_data = {}
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            sheets_data[sheet_name] = {"columns": [], "data": [], "row_count": 0}
            continue

        # First row is header
        headers = [str(h) if h is not None else f"Column_{i}" for i, h in enumerate(rows[0])]
        data = []
        for row in rows[1:]:
            row_dict = {}
            for i, val in enumerate(row):
                col_name = headers[i] if i < len(headers) else f"Column_{i}"
                # Convert to JSON-safe types
                if val is None:
                    row_dict[col_name] = None
                elif isinstance(val, datetime):
                    row_dict[col_name] = val.isoformat()
                elif isinstance(val, (int, float, bool)):
                    row_dict[col_name] = val
                else:
                    row_dict[col_name] = str(val)
            data.append(row_dict)

        sheets_data[sheet_name] = {
            "columns": headers,
            "data": data,
            "row_count": len(data),
        }
        print(f"  Sheet '{sheet_name}': {len(headers)} columns, {len(data)} rows")

    wb.close()
    return sheet_names, sheets_data


def validate_config(sheet_names, sheets_data):
    """
    Validate the config across all sheets.
    Returns a list of validation error dicts.

    ========================================================
    ADD YOUR CUSTOM VALIDATION LOGIC HERE.
    Each error should be a dict with these keys:
        - Sheet: name of the sheet containing the error
        - Row: 1-based row number (relative to data, not header)
        - Column: column name where the error was found
        - Error_Type: category (MISSING_VALUE, INVALID_REFERENCE, DUPLICATE, etc.)
        - Severity: ERROR or WARNING
        - Error_Message: human-readable description
        - Expected_Value: what was expected (optional)
        - Actual_Value: what was found (optional)
    ========================================================
    """
    errors = []
    print("Running validation checks...")

    def add_error(sheet, row, col, error_type, severity, message, expected="", actual=""):
        errors.append({
            "Sheet": sheet,
            "Row": row,
            "Column": col,
            "Error_Type": error_type,
            "Severity": severity,
            "Error_Message": message,
            "Expected_Value": expected,
            "Actual_Value": actual,
        })

    # --- Define required columns per sheet ---
    # Map sheet name -> list of columns that must not be empty
    required_columns = {
        "InputFiles": ["SourceName", "FilePath", "FileType", "Encoding", "Active"],
        "Rules": ["RuleID", "RuleName", "SourceField", "TargetField", "MatchType", "Active"],
        "Enrichment": ["EnrichmentID", "SourceSheet", "LookupField", "LookupSource", "OutputField", "Active"],
        "FieldOperations": ["OperationID", "InputField", "Operation", "OutputField", "Active"],
        "OutputRules": ["OutputID", "OutputName", "OutputPath", "Format", "Active"],
    }

    for sheet_name in sheet_names:
        sheet_info = sheets_data.get(sheet_name, {})
        data = sheet_info.get("data", [])
        columns = sheet_info.get("columns", [])

        if not data:
            continue

        req_cols = required_columns.get(sheet_name, [])

        # Rule 1: Check required columns for empty/null values (ERROR severity)
        for row_idx, row in enumerate(data):
            for col in req_cols:
                if col not in columns:
                    continue
                val = row.get(col)
                if val is None or (isinstance(val, str) and val.strip() == ""):
                    add_error(
                        sheet_name, row_idx + 2, col,
                        "MISSING_VALUE", "ERROR",
                        f"Required field '{col}' is empty",
                        "Non-empty value",
                        "(null)" if val is None else "(empty)",
                    )

        # Rule 2: Check optional columns for empty/null values (WARNING severity)
        for row_idx, row in enumerate(data):
            for col in columns:
                if col in req_cols:
                    continue  # Already checked above
                val = row.get(col)
                if val is None or (isinstance(val, str) and val.strip() == ""):
                    add_error(
                        sheet_name, row_idx + 2, col,
                        "MISSING_VALUE", "WARNING",
                        f"Optional field '{col}' is empty",
                        "Non-empty value",
                        "(null)" if val is None else "(empty)",
                    )

    # Rule 3: Cross-sheet reference validation (Enrichment.SourceSheet must be a valid sheet name)
    enrichment = sheets_data.get("Enrichment", {})
    enr_data = enrichment.get("data", [])
    enr_cols = enrichment.get("columns", [])
    if "SourceSheet" in enr_cols:
        for row_idx, row in enumerate(enr_data):
            ref = row.get("SourceSheet")
            if ref and isinstance(ref, str) and ref.strip() and ref not in sheet_names:
                add_error(
                    "Enrichment", row_idx + 2, "SourceSheet",
                    "INVALID_REFERENCE", "ERROR",
                    f"SourceSheet '{ref}' does not exist in workbook",
                    f"One of: {', '.join(sheet_names)}",
                    ref,
                )

    # Rule 4: Duplicate detection in Rules sheet (same SourceField + TargetField)
    rules = sheets_data.get("Rules", {})
    rules_data = rules.get("data", [])
    rules_cols = rules.get("columns", [])
    if "SourceField" in rules_cols and "TargetField" in rules_cols:
        seen = {}
        for row_idx, row in enumerate(rules_data):
            src = row.get("SourceField", "")
            tgt = row.get("TargetField", "")
            if src and tgt:
                key = f"{src}|{tgt}"
                if key in seen:
                    add_error(
                        "Rules", row_idx + 2, "SourceField",
                        "DUPLICATE", "ERROR",
                        f"Duplicate rule: SourceField='{src}' + TargetField='{tgt}' already defined at row {seen[key]}",
                        "Unique SourceField+TargetField combination",
                        f"Duplicate of row {seen[key]}",
                    )
                else:
                    seen[key] = row_idx + 2

    print(f"Validation complete. Found {len(errors)} issues.")
    return errors


def build_summary(errors, sheet_names):
    """Build validation summary statistics."""
    total_errors = sum(1 for e in errors if e["Severity"] == "ERROR")
    total_warnings = sum(1 for e in errors if e["Severity"] == "WARNING")

    errors_by_sheet = {}
    for e in errors:
        sheet = e["Sheet"]
        errors_by_sheet[sheet] = errors_by_sheet.get(sheet, 0) + 1

    errors_by_type = {}
    for e in errors:
        etype = e["Error_Type"]
        errors_by_type[etype] = errors_by_type.get(etype, 0) + 1

    errors_by_severity = {}
    for e in errors:
        sev = e["Severity"]
        errors_by_severity[sev] = errors_by_severity.get(sev, 0) + 1

    return {
        "total_errors": total_errors,
        "total_warnings": total_warnings,
        "total_issues": len(errors),
        "errors_by_sheet": errors_by_sheet,
        "errors_by_type": errors_by_type,
        "errors_by_severity": errors_by_severity,
        "sheets_validated": sheet_names,
        "validation_duration_seconds": 0,  # Updated below
    }


def write_validated_excel(original_path, errors):
    """
    Create a copy of the Excel file with an added 'Validation' sheet containing errors.
    Returns the path to the validated file.
    """
    if not errors:
        return None

    original = Path(original_path)
    validated_path = original.parent / f"{original.stem}_validated{original.suffix}"
    print(f"Writing validated Excel to: {validated_path}")

    wb = openpyxl.load_workbook(original_path)

    # Remove existing Validation sheet if present
    if "Validation" in wb.sheetnames:
        del wb["Validation"]

    ws = wb.create_sheet("Validation")

    # Write header
    validation_columns = [
        "Sheet", "Row", "Column", "Error_Type", "Severity",
        "Error_Message", "Expected_Value", "Actual_Value"
    ]
    for col_idx, col_name in enumerate(validation_columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = openpyxl.styles.Font(bold=True)

    # Write error rows
    for row_idx, error in enumerate(errors, 2):
        for col_idx, col_name in enumerate(validation_columns, 1):
            val = error.get(col_name, "")
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-fit column widths (approximate)
    for col_idx, col_name in enumerate(validation_columns, 1):
        max_len = len(col_name)
        for row_idx in range(2, min(len(errors) + 2, 102)):  # Check first 100 rows
            cell_val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_len = max(max_len, len(cell_val))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 50)

    wb.save(validated_path)
    wb.close()
    print(f"Validated Excel written with {len(errors)} error rows")
    return str(validated_path)


def main():
    start_time = datetime.now()
    task_id = get_env("TASK_ID", "unknown")
    file_path = get_env("FILE_PATH")
    results_dir = get_env("RESULTS_DIR")

    print("=" * 70)
    print(f"Config Validator Script")
    print(f"Task ID: {task_id}")
    print(f"File: {file_path}")
    print(f"Results Dir: {results_dir}")
    print(f"Started: {start_time.isoformat()}")
    print("=" * 70)

    if not file_path:
        print("ERROR: FILE_PATH environment variable is required")
        sys.exit(1)

    if not Path(file_path).exists():
        print(f"ERROR: File not found: {file_path}")
        sys.exit(1)

    try:
        # Step 1: Read Excel
        print("\n[Step 1/4] Reading Excel file...")
        sheet_names, sheets_data = read_excel_sheets(file_path)

        # Step 2: Validate
        print("\n[Step 2/4] Running validation...")
        errors = validate_config(sheet_names, sheets_data)

        # Step 3: Build summary
        print("\n[Step 3/4] Building summary...")
        validation_passed = len([e for e in errors if e["Severity"] == "ERROR"]) == 0
        summary = build_summary(errors, sheet_names)

        end_time = datetime.now()
        duration = (end_time - start_time).total_seconds()
        summary["validation_duration_seconds"] = round(duration, 2)

        # Step 4: Write outputs
        print("\n[Step 4/4] Writing outputs...")

        # Write validated Excel (only if errors exist)
        validated_file_path = None
        if errors:
            validated_file_path = write_validated_excel(file_path, errors)

        # Build validation sheet data for results JSON
        validation_sheet = None
        if errors:
            validation_columns = [
                "Sheet", "Row", "Column", "Error_Type", "Severity",
                "Error_Message", "Expected_Value", "Actual_Value"
            ]
            validation_sheet = {
                "columns": validation_columns,
                "data": errors,
                "row_count": len(errors),
                "error_count": len(errors),
            }

        # Add error_count per sheet
        for sheet_name in sheet_names:
            sheet_errors = sum(1 for e in errors if e["Sheet"] == sheet_name)
            sheets_data[sheet_name]["error_count"] = sheet_errors

        # Build results JSON
        results = {
            "status": "completed",
            "file_path": file_path,
            "validation_passed": validation_passed,
            "summary": summary,
            "sheet_names": sheet_names,
            "sheets": sheets_data,
            "completed_at": end_time.isoformat(),
        }

        # Add validation sheet to results if errors exist
        if validation_sheet:
            results["sheets"]["Validation"] = validation_sheet
            results["validated_file_path"] = validated_file_path

        # Write results JSON
        if results_dir:
            results_path = Path(results_dir) / f"{task_id}.json"
            results_path.parent.mkdir(parents=True, exist_ok=True)
            with open(results_path, "w", encoding="utf-8") as f:
                json.dump(results, f, indent=2, default=str)
            print(f"Results written to: {results_path}")

        # Print summary
        print("\n" + "=" * 70)
        print(f"VALIDATION {'PASSED' if validation_passed else 'FAILED'}")
        print(f"Total Issues: {len(errors)}")
        print(f"  Errors: {summary['total_errors']}")
        print(f"  Warnings: {summary['total_warnings']}")
        if summary["errors_by_sheet"]:
            print(f"  By Sheet: {summary['errors_by_sheet']}")
        if summary["errors_by_type"]:
            print(f"  By Type: {summary['errors_by_type']}")
        print(f"Duration: {duration:.2f}s")
        if validated_file_path:
            print(f"Validated Excel: {validated_file_path}")
        print("=" * 70)

    except Exception as e:
        print(f"\nFATAL ERROR: {e}")
        traceback.print_exc()

        # Write error results
        if results_dir:
            error_results = {
                "status": "failed",
                "file_path": file_path,
                "validation_passed": None,
                "error": str(e),
                "traceback": traceback.format_exc(),
                "completed_at": datetime.now().isoformat(),
            }
            results_path = Path(results_dir) / f"{task_id}.json"
            results_path.parent.mkdir(parents=True, exist_ok=True)
            with open(results_path, "w", encoding="utf-8") as f:
                json.dump(error_results, f, indent=2)

        sys.exit(1)


if __name__ == "__main__":
    main()
